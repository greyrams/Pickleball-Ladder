# -*- coding: utf-8 -*-
"""
Spyder Editor


"""

# import packages

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Set League Name

league = "Winter2021Mon"

# Change working directory

os.chdir("C:\\Users\\srams\OneDrive\Documents\PickleballLadder\Winter2021Mon")
os.listdir('.')

# Get ranks from last week

initial_file = "WkByWkResults_"+league+".xlsx"
xls_initial = pd.ExcelFile(initial_file)

initial_df = xls_initial.parse('Stats')

WkPos = int(initial_df["Week"].max())

# Load attendance 

Attend = "Attendance_"+league+".xlsx"
xls_attend = pd.ExcelFile(Attend)

attend_sheet = xls_attend.parse(xls_attend.sheet_names[WkPos])

# Merge attendance with starting ranks

LastWk = initial_df[initial_df['Week'] == WkPos][['Player','EndRank','EndGrp']]
LastWk = LastWk.rename(columns = {'EndRank':'StartRank', 'EndGrp':'StartGrp'})

CA = pd.merge(LastWk, attend_sheet, on = 'Player')

attend_sheet.head()

# Rank players who are present, excluding CT court

CA.sort_values('StartRank')

rows = CA.shape[0]

for i in range(0, rows) :
    if (CA.loc[i,'Absent']=="No" and CA.loc[i,'Sub']!="CT") or (CA.loc[i,'Absent']=="Yes" and (CA.loc[i,'Sub']!="No" and CA.loc[i,'Sub']!="CT")) :
        CA.loc[i,'XRank'] = CA.loc[i,'StartRank']

CA['WkRank'] = CA['XRank'].rank()

# Assign court for the week

CA['WkGroup'] = np.ceil(CA['WkRank'].values/4)

for j in range(0, rows) :
    if CA.loc[j,'WkGroup'] >= 1 :
        CA.loc[j, 'WkCourt'] = "Court "+str(round(CA.loc[j,'WkGroup']))
    elif CA.loc[j,'Sub'] == "CT" :
        "Court CT"

CA['WkCourt'] = CA['WkCourt'].str.replace('.0','', regex = False)

CourtIt1 = CA[['Player','WkGroup','WkCourt','Absent','Sub']]
CourtAssign = CourtIt1[pd.notna(CourtIt1['WkCourt'])==True]

# Assign week number, used for tab names in exported Excel files

CurrentWk = WkPos + 1

# Export court assignment to excel

CourtFile = "CourtAssign_"+league+".xlsx"
CourtSheet = "Week"+str(CurrentWk)

CourtAssign.to_excel(CourtFile, sheet_name=CourtSheet, index = False)

# Open results file and export new results sheet to excel

ResultName = "Results_"+league+".xlsx"

writer_res = pd.ExcelWriter(ResultName, engine='openpyxl')
writer_res.book = load_workbook(ResultName)
writer_res.sheets = dict((ws.title, ws) for ws in writer_res.book.worksheets)

ResultSheet = "Week"+str(CurrentWk)

ResultTemp = CourtAssign
ResultTemp['Score'] = ''
ResultTemp = ResultTemp.rename(columns = {'WkGroup' : 'PlayGrp'})

try :
    ResSheet = writer_res.book[sheets[ResultSheet]]
    writer_res.sheets[ResultSheet].delete_rows(1,50)
    ResultTemp.to_excel(writer, sheet_name = ResultSheet, startrow = 1, index = False, header = False)
except NameError as error :
    ResultTemp.to_excel(writer_res, sheet_name = ResultSheet, index = False)
   
writer_res.close()

# Create temporary file for current week

CurrentTemp = CA[['Player','StartRank','StartGrp','Absent','Sub']]

FileTemp = "WeekTemp.csv"
CurrentTemp.to_csv(FileTemp, index = False)

# Format Players for Board Sheets

Boards = CourtAssign[['Player','Absent','Sub','WkGroup']]
assign_rows = Boards.shape[0]

Boards = Boards.reset_index(drop=True)

Boards['BoardPlayer'] = Boards['Player']

for x in range(0,assign_rows) :
    if Boards.loc[x,'Absent'] == "Yes" :
        Boards.loc[x,'BoardPlayer'] = Boards.loc[x,'Sub']
    else :
        Boards.loc[x,'BoardPlayer'] = Boards.loc[x,'Player']

Boards['SpacePos'] = Boards['BoardPlayer'].str.find(" ")
Boards['FirstLastInit'] = Boards['BoardPlayer'].str.extract(pat = '(\D+\s\D)')

BoardList = Boards[['FirstLastInit','WkGroup','Absent']]
RankList = CA[['StartRank','Player','Absent','Sub']]

# Get date for Board Sheet
Dates = "Dates_"+league+".xlsx"
xls_Dates = pd.ExcelFile(Dates)
df_Dates = xls_Dates.parse('Dates')

maxdate = df_Dates['Date'].max()
datewk = df_Dates['Date'][WkPos]
datestr = datewk.strftime('%m/%d/%y')
datedf = pd.DataFrame([datestr])

# Open Board Sheets file, add data, and close

BoardSheets = "BoardSheets_"+league+".xlsx"
writer = pd.ExcelWriter(BoardSheets, engine='openpyxl')
writer.book = load_workbook(BoardSheets)
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

reader = pd.read_excel(r"BoardSheets_"+league+".xlsx")

writer.sheets['Data'].delete_rows(6,56)
datedf.to_excel(writer, sheet_name ='Data', startrow = 2, startcol = 1, index = False, header = False)
BoardList.to_excel(writer, sheet_name ='Data', startrow = 5, index = False, header = False)
RankList.to_excel(writer, sheet_name ='Data', startrow = 5, startcol = 4, index = False, header = False)

writer.close()
